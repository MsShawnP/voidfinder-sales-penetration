"""Broker-ready workbook export.

Two tabs. "Void Summary" carries the verified headline figures;
"Broker Work List" is the field-rep sheet: store numbers, addresses,
what to check, and the dollar figure per line, pre-sorted by
opportunity. The summary total is written FROM the same frame that
fills the detail tab, so the two always reconcile (the verified-
figures rule from the trade-spend diagnostic).
"""

from io import BytesIO

import pandas as pd
from dash import Input, Output, callback, dcc
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app import calculations, data
from app import workbook_styles as st
from app.filters import apply_display_filters, parse_state

_VOID_TYPE_LABEL = {"never_scanned": "Never scanned", "went_dark": "Went dark"}


def generate_workbook(voids: pd.DataFrame, addresses: pd.DataFrame, as_of, params) -> bytes:
    """Build the two-tab workbook and return its bytes."""
    # Clip to the reporting period before anything reads the frame, so
    # the workbook prints the same total as the screen it was exported
    # from and both tabs still reconcile against each other.
    work = calculations.apply_period(voids, params.get("period_weeks")).copy()
    if not addresses.empty:
        work = work.merge(
            addresses.rename(columns={"state": "addr_state"}),
            on="store_id", how="left",
        )
    else:
        for col in ("street", "city", "addr_state", "zip"):
            work[col] = ""

    wb = Workbook()
    wb.remove(wb.active)
    summary = wb.create_sheet(title=st.TAB_NAMES[0])
    summary.sheet_properties.tabColor = st.CHICAGO_20
    detail = wb.create_sheet(title=st.TAB_NAMES[1])
    detail.sheet_properties.tabColor = st.HK_35

    _build_summary(summary, work, as_of, params)
    _build_detail(detail, work)

    wb.active = 0
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_summary(ws, work, as_of, params):
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [3, 34, 18, 18, 18])

    ws["B2"] = "Cinderhaven Provisions"
    ws["B2"].font = st.FONT_HEADER
    ws["B3"] = "Void Finder — Store Void Opportunity Report"
    ws["B3"].font = st.FONT_BODY
    as_of_str = str(as_of)[:10] if as_of is not None else "—"
    period_label = params.get("period_label") or params.get("period") or "all history"
    ws["B4"] = (
        f"As of week ending {as_of_str} · period {period_label} · "
        f"void threshold {params['void_weeks_n']} weeks · slow-mover floor "
        f"{params['slow_mover_min']} units/week"
    )
    ws["B4"].font = st.FONT_SMALL

    total = work["void_dollars"].sum() if not work.empty else 0.0
    kpis = [
        ("Total void opportunity", f"${total:,.0f}"),
        ("Open voids", f"{len(work):,}"),
        ("Stores affected", f"{work['store_id'].nunique():,}" if not work.empty else "0"),
    ]
    row = 6
    for label, value in kpis:
        ws.cell(row=row, column=2, value=value).font = st.FONT_KPI_VALUE
        ws.cell(row=row + 1, column=2, value=label).font = st.FONT_KPI_LABEL
        row += 3

    ws.cell(row=row, column=2, value="By void type").font = st.FONT_SECTION
    row += 1
    headers = ["Type", "Voids", "Stores", "Opportunity"]
    for col, h in enumerate(headers, start=2):
        c = ws.cell(row=row, column=col, value=h)
        c.font = st.FONT_TABLE_HEADER
        c.fill = st.FILL_HEADER
    row += 1
    if not work.empty:
        by_type = work.groupby("void_type", observed=True).agg(
            voids=("sku", "size"),
            stores=("store_id", "nunique"),
            dollars=("void_dollars", "sum"),
        )
        for vt, r in by_type.iterrows():
            ws.cell(row=row, column=2, value=_VOID_TYPE_LABEL.get(vt, vt)).font = st.FONT_BODY
            ws.cell(row=row, column=3, value=int(r["voids"])).font = st.FONT_BODY
            ws.cell(row=row, column=4, value=int(r["stores"])).font = st.FONT_BODY
            d = ws.cell(row=row, column=5, value=float(r["dollars"]))
            d.font = st.FONT_BODY
            d.number_format = st.NUM_FMT_DOLLAR
            row += 1

    row += 1
    ws.cell(
        row=row, column=2,
        value=(
            "Opportunity = median weekly dollars of comparable scanning "
            "stores (same volume tier + region) × weeks without a scan. "
            "Median, not mean. Items whose comparables sell below the "
            "slow-mover floor are excluded. Dollars count only the void "
            "weeks that fall inside the reporting period stated above. "
            "Summary figures are computed from the same rows as the "
            "Broker Work List tab."
        ),
    ).font = st.FONT_SMALL


_DETAIL_COLUMNS = [
    ("chain_name", "Retailer", 13),
    ("store_id", "Store #", 17),
    ("street", "Street", 26),
    ("city", "City", 16),
    ("addr_state", "State", 7),
    ("zip", "ZIP", 9),
    ("region", "Region", 11),
    ("volume_tier", "Tier", 8),
    ("sku", "SKU", 13),
    ("product_name", "Item", 30),
    ("void_type_label", "Void type", 14),
    ("void_weeks", "Weeks dark", 11),
    ("void_dollars", "Opportunity ($)", 15),
    ("fixability", "Fixability", 10),
    ("priority", "Priority ($)", 13),
    ("action", "What to check", 44),
]


def _action_for(row) -> str:
    if row["void_type"] == "never_scanned":
        if pd.notna(row.get("cluster_id")):
            return (
                "Part of a regional reset failure — confirm the item was "
                "set to planogram; one reset call covers the whole cluster."
            )
        return "Authorized but never set — confirm shelf placement and tag."
    if row["void_weeks"] > 12:
        return "Dark for months — verify item still has a home; re-cut PO if delisted in error."
    return "Was selling, stopped — check for phantom OOS, lost facing, or tag pull."


def _build_detail(ws, work):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for col, (_, header, width) in enumerate(_DETAIL_COLUMNS, start=1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = st.FONT_TABLE_HEADER
        c.fill = st.FILL_HEADER
        ws.column_dimensions[get_column_letter(col)].width = width

    if work.empty:
        return

    rows = work.sort_values("void_dollars", ascending=False).copy()
    rows["void_type_label"] = rows["void_type"].map(_VOID_TYPE_LABEL)
    rows["action"] = rows.apply(_action_for, axis=1)

    for r_idx, (_, row) in enumerate(rows.iterrows(), start=2):
        for c_idx, (field, _, _) in enumerate(_DETAIL_COLUMNS, start=1):
            value = row.get(field, "")
            if pd.isna(value):
                value = ""
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = st.FONT_BODY
            if field in ("void_dollars", "priority"):
                cell.number_format = st.NUM_FMT_DOLLAR
            if row["void_type"] == "never_scanned" and field == "void_type_label":
                cell.fill = st.FILL_BAD


def register_export_callback():
    @callback(
        Output("download-workbook", "data"),
        Input("btn-download-workbook", "n_clicks"),
        Input("filter-state", "data"),
        prevent_initial_call=True,
    )
    def _download(n_clicks, filter_json):
        from dash import ctx

        if ctx.triggered_id != "btn-download-workbook" or not n_clicks:
            return None
        state = parse_state(filter_json)
        voids = data.get_voids(state["void_weeks_n"], state["slow_mover_min"])
        shown = apply_display_filters(voids, state)
        # The period is resolved here, where the week grid is available;
        # generate_workbook stays pure so the export is unit-testable.
        window = data.period_window(
            state["period"], state.get("custom_start"), state.get("custom_end"),
        )
        params = dict(state)
        if window:
            params["period_weeks"] = window["period_weeks"]
            params["period_label"] = window["label"]
        payload = generate_workbook(shown, data.get_addresses(), data.as_of_week(), params)
        return dcc.send_bytes(payload, "voidfinder-broker-work-list.xlsx")
