"""Workbook export tests — the verified-figures rule: summary tab
totals must be computed from the same rows as the detail tab."""

from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.export import generate_workbook

PARAMS = {"void_weeks_n": 6, "slow_mover_min": 0.5}

# The real caller resolves the selected period against the week grid and passes
# the result down (export.py register_export_callback). PARAMS above omits it,
# so those tests exercise the no-period path; these exercise the clip.
PARAMS_WITH_PERIOD = {
    **PARAMS, "period": "26w", "period_weeks": 26,
    "period_label": "the last 26 weeks",
}


def _voids():
    """Two voids whose lives straddle a 26-week window.

    ``median_weekly_dollars`` is what ``calculations.period_void_dollars``
    clips against; ``void_dollars`` is the unclipped whole-life figure and
    equals weekly x weeks for each row.
    """
    return pd.DataFrame(
        [
            {
                "sku": "CHP-AS-001", "store_id": "RET-KROGER-S0001",
                "chain_name": "Kroger", "region": "Southeast",
                "volume_tier": "medium", "product_name": "Basil Marinara",
                "void_type": "never_scanned", "void_weeks": 40,
                "median_weekly_dollars": 20.0,
                "void_dollars": 800.0, "fixability": 0.95, "priority": 380.0,
                "cluster_id": "RET-KROGER|Southeast",
            },
            {
                "sku": "CHP-SC-002", "store_id": "RET-WALMART-S0002",
                "chain_name": "Walmart", "region": "West",
                "volume_tier": "high", "product_name": "Hot Honey Mustard",
                "void_type": "went_dark", "void_weeks": 8,
                "median_weekly_dollars": 18.75,
                "void_dollars": 150.0, "fixability": 0.7, "priority": 105.0,
                "cluster_id": None,
            },
        ]
    )


def _addresses():
    return pd.DataFrame(
        [
            {"store_id": "RET-KROGER-S0001", "street": "100 Peach St",
             "city": "Atlanta", "state": "GA", "zip": "30301"},
            {"store_id": "RET-WALMART-S0002", "street": "8 Desert Rd",
             "city": "Phoenix", "state": "AZ", "zip": "85001"},
        ]
    )


def _load(payload):
    return load_workbook(BytesIO(payload))


def test_workbook_has_two_branded_tabs():
    wb = _load(generate_workbook(_voids(), _addresses(), "2025-12-27", PARAMS))
    assert wb.sheetnames == ["Void Summary", "Broker Work List"]


def test_summary_total_reconciles_with_detail_rows():
    voids = _voids()
    wb = _load(generate_workbook(voids, _addresses(), "2025-12-27", PARAMS))
    detail = wb["Broker Work List"]
    header = [c.value for c in detail[1]]
    dollars_col = header.index("Opportunity ($)") + 1
    detail_total = sum(
        detail.cell(row=r, column=dollars_col).value
        for r in range(2, detail.max_row + 1)
    )
    assert detail_total == pytest.approx(voids["void_dollars"].sum())


def test_detail_rows_sorted_by_opportunity_and_carry_addresses():
    wb = _load(generate_workbook(_voids(), _addresses(), "2025-12-27", PARAMS))
    detail = wb["Broker Work List"]
    header = [c.value for c in detail[1]]
    assert "Street" in header and "ZIP" in header and "What to check" in header
    dollars_col = header.index("Opportunity ($)") + 1
    values = [
        detail.cell(row=r, column=dollars_col).value
        for r in range(2, detail.max_row + 1)
    ]
    assert values == sorted(values, reverse=True)
    street_col = header.index("Street") + 1
    assert detail.cell(row=2, column=street_col).value == "100 Peach St"


def test_missing_addresses_do_not_break_export():
    wb = _load(generate_workbook(_voids(), pd.DataFrame(), "2025-12-27", PARAMS))
    detail = wb["Broker Work List"]
    assert detail.max_row == 3  # header + 2 rows, blanks not crashes


def test_empty_voids_produce_valid_workbook():
    empty = _voids().iloc[0:0]
    wb = _load(generate_workbook(empty, _addresses(), None, PARAMS))
    assert wb["Broker Work List"].max_row == 1  # header only


# ── The workbook total is period-clipped ────────────────────────────────────
# generate_workbook clips the frame once, before either tab reads it, so the
# export prints the same total as the screen it came from and the two tabs
# still reconcile. Fixed 2026-07-28; these were strict-xfail until then.

def _summary_dollar_cells(wb):
    return [
        c.value for row in wb["Void Summary"].iter_rows() for c in row
        if isinstance(c.value, str) and c.value.startswith("$")
    ]


def test_summary_total_is_clipped_to_the_selected_period():
    voids = _voids()
    wb = _load(generate_workbook(voids, _addresses(), "2025-12-27", PARAMS_WITH_PERIOD))
    unclipped = voids["void_dollars"].sum()
    assert f"${unclipped:,.0f}" not in _summary_dollar_cells(wb), (
        "summary KPI still shows the unclipped whole-life total"
    )


def test_summary_and_detail_still_reconcile_under_a_period():
    # The verified-figures rule has to survive the clip: both tabs read
    # one already-clipped frame, so a period never splits them apart.
    wb = _load(generate_workbook(_voids(), _addresses(), "2025-12-27", PARAMS_WITH_PERIOD))
    detail = wb["Broker Work List"]
    header = [c.value for c in detail[1]]
    dollars_col = header.index("Opportunity ($)") + 1
    detail_total = sum(
        detail.cell(row=r, column=dollars_col).value
        for r in range(2, detail.max_row + 1)
    )
    # 40-week void clipped to 26 x $20 = $520; 8-week void unchanged = $150.
    assert detail_total == pytest.approx(670.0)
    assert f"${detail_total:,.0f}" in _summary_dollar_cells(wb)


def test_parameter_line_states_the_reporting_period():
    wb = _load(generate_workbook(_voids(), _addresses(), "2025-12-27", PARAMS_WITH_PERIOD))
    text = " ".join(
        str(c.value) for row in wb["Void Summary"].iter_rows() for c in row
        if isinstance(c.value, str)
    )
    assert "26w" in text or "26 week" in text, (
        "reader cannot tell which period the total covers"
    )
